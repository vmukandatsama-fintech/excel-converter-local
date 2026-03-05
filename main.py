import time
import os
import sys
import json
import re
from copy import copy
import threading
import datetime
import subprocess
import smtplib
import ssl
import shutil
import configparser
import ctypes
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication

# --- THIRD PARTY IMPORTS ---
try:
    import win32com.client
    import pythoncom
    import openpyxl
    from watchdog.observers import Observer
    from watchdog.events import PatternMatchingEventHandler
    from pystray import Icon, Menu, MenuItem
    from PIL import Image, ImageDraw
    from winotify import Notification, audio
except ImportError as e:
    print(f"Missing dependencies: {e}")
    print("Run: pip install winotify")
    sys.exit(1)

def get_base_path():
    if getattr(sys, 'frozen', False): return Path(sys.executable).parent
    return Path(__file__).resolve().parent

def log_message(message):
    try:
        base_path = get_base_path()
        with open(base_path / "activity_log.txt", "a") as f:
            f.write(f"[{datetime.datetime.now()}] {message}\n")
    except: pass
    console_log(message)

CONSOLE_LOCK = threading.Lock()

def console_log(message):
    try:
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with CONSOLE_LOCK:
            print(f"[{timestamp}] {message}")
    except Exception:
        pass

# --- DYNAMIC PATH LOGIC ---
def get_root_directory():
    """Get root directory from config.ini or fall back to user profile."""
    try:
        config_path = get_base_path() / "config.ini"
        if config_path.exists():
            config = configparser.ConfigParser()
            config.read(config_path)
            if 'PATHS' in config:
                root_path = config['PATHS'].get('root_path', '')
                if root_path:
                    # Support both absolute paths and network paths
                    return Path(root_path)
    except Exception as e:
        console_log(f"Error reading config path: {e}")
    
    # Fallback to user profile
    return Path(os.environ['USERPROFILE']) / "Premium Leaf Zimbabwe" / "v2SavannahTEST - Collection Vouchers"

ROOT_DIR = get_root_directory()
JSON_DIR = ROOT_DIR / "Json"
TEMPLATE_DIR = ROOT_DIR / "Templates"
WORK_DIR = ROOT_DIR / "Populated Template"
OUTPUT_DIR = ROOT_DIR / "PDF CVs"
TEMPLATE_FILE = TEMPLATE_DIR / "Collection Voucher Template.xlsx"
BACKUP_DIR = JSON_DIR / "Backup"

def initialize_folders():
    """Builds the 4-folder structure if it doesn't exist."""
    for folder in [JSON_DIR, TEMPLATE_DIR, WORK_DIR, OUTPUT_DIR, BACKUP_DIR]:
        folder.mkdir(parents=True, exist_ok=True)

def get_versioned_path(base_path):
    """Return a non-conflicting path by appending _01, _02 if needed."""
    if not base_path.exists():
        return base_path
    stem = base_path.stem
    suffix = base_path.suffix
    counter = 1
    while True:
        candidate = base_path.with_name(f"{stem}_{counter:02d}{suffix}")
        if not candidate.exists():
            return candidate
        counter += 1

def archive_processed_json(processing_path):
    """Move processed .processing file to Json/Processed as .json for audit/history."""
    try:
        processing_path = Path(processing_path)
        processed_dir = JSON_DIR / "Processed"
        processed_dir.mkdir(parents=True, exist_ok=True)

        archived_name = f"{processing_path.stem}_processed.json"
        archived_path = get_versioned_path(processed_dir / archived_name)
        os.rename(processing_path, archived_path)
        log_message(f"ARCHIVED JSON: {archived_path.name}")
        return archived_path
    except Exception as e:
        log_message(f"ARCHIVE ERROR: Could not archive {processing_path} - {e}")
        raise

def backup_claimed_json(processing_path):
    """Copy claimed .processing file to Json/Backup as .json before processing."""
    try:
        processing_path = Path(processing_path)
        BACKUP_DIR.mkdir(parents=True, exist_ok=True)

        backup_name = f"{processing_path.stem}_backup.json"
        backup_path = get_versioned_path(BACKUP_DIR / backup_name)
        shutil.copy2(processing_path, backup_path)
        log_message(f"BACKUP JSON: {backup_path.name}")
        return backup_path
    except Exception as e:
        log_message(f"BACKUP ERROR: Could not backup {processing_path} - {e}")
        raise

def normalize_excel_value(value):
    """Convert complex JSON values into scalars supported by openpyxl."""
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool, datetime.date, datetime.datetime)):
        return value
    if isinstance(value, dict):
        for key in ("Value", "value", "Name", "name", "Label", "label", "Text", "text", "Id", "id"):
            if key in value and value[key] is not None:
                return normalize_excel_value(value[key])
        return json.dumps(value, ensure_ascii=False)
    if isinstance(value, list):
        return ", ".join(str(normalize_excel_value(v)) for v in value if v is not None)
    return str(value)

def sanitize_filename(value, default="Unknown"):
    """Create a Windows-safe filename segment from arbitrary input."""
    text = str(value or "").strip()
    if not text:
        text = default
    text = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text)
    text = text.rstrip(". ")
    return text if text else default

def is_archival_json_name(file_name):
    """Return True for JSON artifacts that should never be re-processed."""
    lower_name = str(file_name).lower()
    return lower_name.endswith("_processed.json") or "_backup" in lower_name or "_failed" in lower_name

def safe_set_cell(ws, cell_ref, raw_value, field_name):
    """Normalize and safely assign a value to an Excel cell."""
    value = normalize_excel_value(raw_value)
    try:
        ws[cell_ref] = value
        force_cell_font_black(ws[cell_ref])
    except Exception as e:
        log_message(f"EXCEL ASSIGN ERROR [{field_name} @ {cell_ref}]: raw={raw_value!r} normalized={value!r} err={e}")
        raise

def safe_set_cell_rc(ws, row, column, raw_value, field_name):
    """Normalize and safely assign a value using row/column coordinates."""
    value = normalize_excel_value(raw_value)
    try:
        cell = ws.cell(row=row, column=column)
        cell.value = value
        force_cell_font_black(cell)
    except Exception as e:
        log_message(f"EXCEL ASSIGN ERROR [{field_name} @ r{row}c{column}]: raw={raw_value!r} normalized={value!r} err={e}")
        raise

def force_cell_font_black(cell):
    """Ensure populated values remain visible regardless of template/theme colors."""
    try:
        updated_font = copy(cell.font)
        updated_font.color = "000000"
        cell.font = updated_font
    except Exception as e:
        log_message(f"EXCEL FONT WARNING: Could not force black font at {cell.coordinate} - {e}")

class EmailSender:
    """Handles email sending functionality for PDF distribution."""
    
    def __init__(self, config_path=None):
        self.enabled = False
        self.smtp_server = None
        self.smtp_port = 587
        self.sender_email = None
        self.sender_password = None
        self.recipients = []
        self.subject = "New Collection Voucher PDF Generated"
        self.body_template = "Please find the attached PDF."
        
        if config_path is None:
            config_path = get_base_path() / "config.ini"
        
        self.load_config(config_path)
    
    def load_config(self, config_path):
        """Load email configuration from config.ini file."""
        try:
            if not Path(config_path).exists():
                log_message("Config file not found, email disabled")
                return
            
            config = configparser.ConfigParser()
            config.read(config_path)
            
            if 'EMAIL' not in config:
                log_message("No EMAIL section in config, email disabled")
                return
            
            email_config = config['EMAIL']
            
            # Parse boolean value
            self.enabled = email_config.get('enabled', 'false').lower() in ['true', '1', 'yes']
            
            if not self.enabled:
                log_message("Email notifications are disabled in config")
                return
            
            self.smtp_server = email_config.get('smtp_server', '')
            self.smtp_port = int(email_config.get('smtp_port', '587'))
            self.sender_email = email_config.get('sender_email', '')
            self.sender_password = email_config.get('sender_password', '')
            
            # Parse recipients (comma-separated)
            recipients_str = email_config.get('recipients', '')
            self.recipients = [r.strip() for r in recipients_str.split(',') if r.strip()]
            
            self.subject = email_config.get('subject', self.subject)
            self.body_template = email_config.get('body_template', self.body_template)
            
            # Validate configuration
            if not all([self.smtp_server, self.sender_email, self.sender_password, self.recipients]):
                log_message("EMAIL: Incomplete configuration, email disabled")
                self.enabled = False
            else:
                log_message(f"EMAIL: Enabled - Will send to {len(self.recipients)} recipient(s)")
                
        except Exception as e:
            log_message(f"EMAIL: Config error - {e}")
            self.enabled = False
    
    def send_pdf(self, pdf_path, req_no):
        """Send PDF file via email to configured recipients."""
        if not self.enabled:
            return False
        
        try:
            pdf_path = Path(pdf_path)
            if not pdf_path.exists():
                log_message(f"EMAIL: PDF file not found: {pdf_path}")
                return False
            
            # Create message
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = ', '.join(self.recipients)

            # Format subject/body with template variables
            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            template_values = {
                "req_no": req_no,
                "timestamp": timestamp,
                "filename": pdf_path.name
            }

            try:
                msg['Subject'] = self.subject.format(**template_values)
            except Exception:
                msg['Subject'] = self.subject

            try:
                body = self.body_template.format(**template_values)
            except Exception:
                body = self.body_template
            
            msg.attach(MIMEText(body, 'plain'))
            
            # Attach PDF
            with open(pdf_path, 'rb') as f:
                pdf_attachment = MIMEApplication(f.read(), _subtype='pdf')
                pdf_attachment.add_header('Content-Disposition', 'attachment', 
                                        filename=pdf_path.name)
                msg.attach(pdf_attachment)
            
            # Send email
            log_message(f"EMAIL: Connecting to {self.smtp_server}:{self.smtp_port}")
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)
            
            log_message(f"EMAIL: Successfully sent {pdf_path.name} to {len(self.recipients)} recipient(s)")
            return True
            
        except smtplib.SMTPAuthenticationError:
            log_message("EMAIL ERROR: Authentication failed - check credentials")
            return False
        except (smtplib.SMTPServerDisconnected, ssl.SSLError, TimeoutError) as e:
            log_message(f"EMAIL ERROR: Communication error - {e}")
            return False
        except smtplib.SMTPException as e:
            log_message(f"EMAIL ERROR: SMTP error - {e}")
            return False
        except Exception as e:
            log_message(f"EMAIL ERROR: {e}")
            return False

    def send_failure_notification(self, req_no, source_file, attempts, error_message):
        """Send failure notification when processing retries are exhausted."""
        if not self.enabled:
            return False

        try:
            msg = MIMEMultipart()
            msg['From'] = self.sender_email
            msg['To'] = ', '.join(self.recipients)
            msg['Subject'] = f"FAILED: Collection Voucher {req_no}"

            timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            body = (
                "Collection Voucher processing failed after maximum retries.\n\n"
                f"Request: {req_no}\n"
                f"Source File: {source_file}\n"
                f"Attempts: {attempts}\n"
                f"Time: {timestamp}\n"
                f"Last Error: {error_message}\n"
            )
            msg.attach(MIMEText(body, 'plain'))

            log_message(f"EMAIL: Sending failure notification for {req_no}")
            with smtplib.SMTP(self.smtp_server, self.smtp_port) as server:
                server.starttls()
                server.login(self.sender_email, self.sender_password)
                server.send_message(msg)

            log_message(f"EMAIL: Failure notification sent for {req_no}")
            return True
        except Exception as e:
            log_message(f"EMAIL ERROR: Failed to send failure notification - {e}")
            return False

class SplashScreen:
    def show(self):
        try:
            os.system("cls")
            print("=" * 60)
            print("PLZ CV ENGINE v2 - PRODUCTION READY")
            print(f"Monitoring Path: {ROOT_DIR}")

            print("Mode: LOCAL")
            print("System: Monitoring for local Power Automate JSON syncs...")
            print("=" * 60)
        except: pass

class CVAutomator:
    def __init__(self):
        self.status = "Idle"
        self.status_lock = threading.Lock()
        self.icon = None
        self.observer = None
        self.poll_thread = None
        self.stop_event = threading.Event()
        self.email_sender = EmailSender()
        self.sleep_guard_enabled = False
        self.processing_stale_seconds = 60
        self.max_process_retries = 3
        self.recovery_lock = threading.Lock()
        self.recovery_in_progress = set()
        self.retry_lock = threading.Lock()
        self.retry_counts = {}
        self.active_processing_lock = threading.Lock()
        self.active_processing = set()
        self.export_lock = threading.Lock()
        self.serialize_pdf_export = True
        self.load_runtime_config()

    def should_skip_json_artifact(self, file_path):
        file_name = Path(file_path).name
        if is_archival_json_name(file_name):
            log_message(f"SKIPPED ARTIFACT: Ignoring archival JSON {file_name}")
            return True
        return False

    def try_mark_processing_active(self, processing_path):
        processing_key = str(Path(processing_path))
        with self.active_processing_lock:
            if processing_key in self.active_processing:
                return False
            self.active_processing.add(processing_key)
            return True

    def unmark_processing_active(self, processing_path):
        processing_key = str(Path(processing_path))
        with self.active_processing_lock:
            self.active_processing.discard(processing_key)

    def is_processing_active(self, processing_path):
        processing_key = str(Path(processing_path))
        with self.active_processing_lock:
            return processing_key in self.active_processing

    def touch_processing_file(self, processing_path):
        """Refresh .processing mtime so stale-recovery logic doesn't race active work."""
        try:
            processing_path = Path(processing_path)
            if processing_path.exists():
                os.utime(processing_path, None)
        except Exception as e:
            log_message(f"TOUCH WARNING: Could not refresh {Path(processing_path).name} - {e}")

    def apply_excel_page_setup(self, worksheet_api):
        """Best-effort PageSetup tuning; some systems reject specific properties."""
        try:
            page_setup = worksheet_api.PageSetup
        except Exception as e:
            log_message(f"EXPORT WARNING: Could not access PageSetup - {e}")
            return

        for name, value in (
            ("PaperSize", 9),  # xlPaperA4
            ("Orientation", 1),  # xlPortrait
            ("Zoom", False),
            ("FitToPagesWide", 1),
            ("FitToPagesTall", 1),
        ):
            try:
                setattr(page_setup, name, value)
            except Exception as e:
                log_message(f"EXPORT WARNING: Could not set PageSetup.{name}={value!r} - {e}")

    def load_runtime_config(self):
        """Load runtime tuning values from config.ini [APP] section."""
        try:
            config_path = get_base_path() / "config.ini"
            if not config_path.exists():
                return

            config = configparser.ConfigParser()
            config.read(config_path)
            if 'APP' not in config:
                return

            app_config = config['APP']
            stale_seconds = app_config.getint('stale_processing_seconds', fallback=self.processing_stale_seconds)
            max_retries = app_config.getint('max_process_retries', fallback=self.max_process_retries)
            self.serialize_pdf_export = app_config.getboolean('serialize_pdf_export', fallback=self.serialize_pdf_export)

            self.processing_stale_seconds = max(15, stale_seconds)
            self.max_process_retries = max(1, max_retries)
            log_message(
                f"APP CONFIG: stale_processing_seconds={self.processing_stale_seconds}, "
                f"max_process_retries={self.max_process_retries}, "
                f"serialize_pdf_export={self.serialize_pdf_export}"
            )
        except Exception as e:
            log_message(f"APP CONFIG ERROR: {e} (using defaults)")

    def get_retry_count(self, processing_path):
        key = str(Path(processing_path).stem)
        with self.retry_lock:
            return self.retry_counts.get(key, 0)

    def increment_retry_count(self, processing_path):
        key = str(Path(processing_path).stem)
        with self.retry_lock:
            self.retry_counts[key] = self.retry_counts.get(key, 0) + 1
            return self.retry_counts[key]

    def clear_retry_count(self, processing_path):
        key = str(Path(processing_path).stem)
        with self.retry_lock:
            self.retry_counts.pop(key, None)

    def move_to_failed(self, processing_path, error_message, attempts):
        processing_path = Path(processing_path)
        failed_dir = JSON_DIR / "Failed"
        failed_dir.mkdir(parents=True, exist_ok=True)

        failed_name = f"{processing_path.stem}_failed.json"
        failed_path = get_versioned_path(failed_dir / failed_name)

        os.rename(processing_path, failed_path)
        log_message(f"FAILED JSON: Moved to {failed_path.name} after {attempts} attempts")

        req_no = sanitize_filename(processing_path.stem, default="Unknown")
        self.email_sender.send_failure_notification(req_no, processing_path.name, attempts, str(error_message))
        self.clear_retry_count(processing_path)

    def restore_failed_json_for_retry(self, failed_path):
        """Copy a failed JSON artifact back into Json/ as a retry candidate."""
        failed_path = Path(failed_path)
        stem = failed_path.stem
        if stem.lower().endswith("_failed"):
            stem = stem[:-7]
        retry_path = get_versioned_path(JSON_DIR / f"{stem}.json")
        shutil.copy2(failed_path, retry_path)
        log_message(f"RETRY QUEUED: {failed_path.name} -> {retry_path.name}")
        return retry_path

    def requeue_recent_failed(self, count=2):
        """Requeue the most recent failed JSON artifacts for automatic retry."""
        try:
            failed_dir = JSON_DIR / "Failed"
            if not failed_dir.exists():
                return

            failed_files = [entry for entry in failed_dir.glob("*_failed.json") if entry.is_file()]
            if not failed_files:
                return

            failed_files.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            selected = failed_files[:max(0, int(count))]
            for failed_file in selected:
                self.restore_failed_json_for_retry(failed_file)

            log_message(f"RETRY BOOTSTRAP: Requeued {len(selected)} recent failed request(s)")
        except Exception as e:
            log_message(f"RETRY BOOTSTRAP ERROR: {e}")

    def prevent_system_sleep(self):
        """Prevent Windows from entering sleep while watcher is active."""
        try:
            ES_CONTINUOUS = 0x80000000
            ES_SYSTEM_REQUIRED = 0x00000001
            result = ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS | ES_SYSTEM_REQUIRED)
            if result == 0:
                log_message("SLEEP GUARD: Failed to enable sleep prevention")
                self.sleep_guard_enabled = False
            else:
                self.sleep_guard_enabled = True
                log_message("SLEEP GUARD: Enabled (system sleep blocked while watcher runs)")
        except Exception as e:
            self.sleep_guard_enabled = False
            log_message(f"SLEEP GUARD ERROR: Could not enable sleep prevention - {e}")

    def restore_system_sleep(self):
        """Restore default Windows sleep behavior."""
        try:
            ES_CONTINUOUS = 0x80000000
            result = ctypes.windll.kernel32.SetThreadExecutionState(ES_CONTINUOUS)
            if result == 0:
                log_message("SLEEP GUARD: Failed to restore default sleep behavior")
            elif self.sleep_guard_enabled:
                log_message("SLEEP GUARD: Disabled (default sleep behavior restored)")
            self.sleep_guard_enabled = False
        except Exception as e:
            log_message(f"SLEEP GUARD ERROR: Could not restore sleep behavior - {e}")

    def set_status(self, message, notify=False):
        with self.status_lock:
            self.status = message
        console_log(f"STATUS: {message}")
        if self.icon:
            self.icon.title = f"PLZ CV Engine - {message}"
            if notify:
                # Use Windows native toast notification instead of pystray
                try:
                    toast = Notification(
                        app_id="PLZ CV Engine",
                        title="PLZ CV Engine",
                        msg=message,
                        duration="short"
                    )
                    toast.show()
                except Exception as e:
                    log_message(f"Notification error: {e}")
                    # Fallback to pystray notification
                    try:
                        self.icon.notify(message, title="PLZ CV Engine")
                    except:
                        pass

    def process_json(self, json_path):
        processing_path = None
        try:
            json_path = Path(json_path)
            if self.should_skip_json_artifact(json_path):
                return
            self.set_status(f"Detected: {json_path.name}")
            
            # --- CLAIM FILE IMMEDIATELY (Prevent multi-PC conflicts) ---
            # Rename file to .processing to claim ownership
            processing_path = json_path.with_suffix('.processing')
            if not self.try_mark_processing_active(processing_path):
                log_message(f"SKIPPED: {json_path.name} (already being processed locally)")
                return
            try:
                # Atomic rename - only one PC will succeed
                os.rename(json_path, processing_path)
                self.touch_processing_file(processing_path)
                log_message(f"CLAIMED: {json_path.name} (this PC will process it)")
            except (FileNotFoundError, PermissionError, OSError) as e:
                # Another PC already claimed it or file doesn't exist
                log_message(f"SKIPPED: {json_path.name} (already claimed by another PC or missing)")
                return

            # Backup first after claim, before parsing/population/conversion
            backup_claimed_json(processing_path)
            
            # --- SAFE READ RETRY LOOP (Fixes 'Line 1 Char 0' error) ---
            data = None
            retries = 3
            while retries > 0:
                if processing_path.exists() and os.path.getsize(processing_path) > 0:
                    try:
                        with open(processing_path, 'r') as f:
                            data = json.load(f)
                        break 
                    except (json.JSONDecodeError, PermissionError):
                        pass # File is still being locked or written by OneDrive
                time.sleep(1)
                retries -= 1

            if data is None:
                log_message(f"FAILED: Sync timeout for {processing_path.name}")
                # Clean up the .processing file
                try:
                    os.remove(processing_path)
                except:
                    pass
                return

            payload = data
            if isinstance(data, dict) and isinstance(data.get("body"), str):
                try:
                    payload = json.loads(data.get("body", "{}"))
                    log_message("Detected wrapped payload format; parsed nested 'body' JSON")
                except json.JSONDecodeError as e:
                    log_message(f"Invalid nested body JSON: {e}")
                    payload = data

            self.touch_processing_file(processing_path)

            req_no_raw = normalize_excel_value(payload.get("FileName") or processing_path.stem)
            req_no = sanitize_filename(req_no_raw, default=processing_path.stem)
            header = payload.get("Header", {})
            lines = payload.get("Lines", [])

            if not TEMPLATE_FILE.exists():
                log_message(f"CRITICAL: Template missing at {TEMPLATE_FILE}")
                return

            # --- POPULATE EXCEL ---
            self.set_status(f"Populating Template for {req_no}...")
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            ws = wb["CollectionVoucher"] if "CollectionVoucher" in wb.sheetnames else wb.active

            safe_set_cell(ws, "D5", f"COLLECTION VOUCHER : {req_no}", "Header.FileName")
            safe_set_cell(ws, "D9", header.get("Requestor"), "Header.Requestor")
            safe_set_cell(ws, "D10", header.get("Date"), "Header.Date")
            safe_set_cell(ws, "D11", header.get("Approval"), "Header.Approval")
            safe_set_cell(ws, "D12", header.get("Authorization"), "Header.Authorization")
            safe_set_cell(ws, "D13", header.get("Comments"), "Header.Comments")
            
            safe_set_cell(ws, "G9", header.get("Driver"), "Header.Driver")
            safe_set_cell(ws, "G10", header.get("DriverID"), "Header.DriverID")
            safe_set_cell(ws, "G11", header.get("Truck"), "Header.Truck")
            safe_set_cell(ws, "G12", header.get("Trailer"), "Header.Trailer")
            safe_set_cell(ws, "G13", header.get("Farmer"), "Header.Farmer")

            for i, item in enumerate(lines):
                row = 18 + i
                line_value = item.get("Line", item.get("line", item.get("LineNumber", item.get("lineNumber"))))
                uom_value = item.get("UOM", item.get("Uom"))
                item_value = item.get("Item", item.get("Description"))
                requested_value = item.get("Requested", item.get("requested"))
                issue_value = item.get("Issue", item.get("ThisIssue", item.get("Issued")))
                already_issued_value = item.get("AlreadyIssued", item.get("TotalToDate"))
                balance_value = item.get("Balance", item.get("Remaining"))

                safe_set_cell_rc(ws, row, 2, line_value, f"Lines[{i}].Line")
                safe_set_cell_rc(ws, row, 3, uom_value, f"Lines[{i}].UOM")
                safe_set_cell_rc(ws, row, 4, item_value, f"Lines[{i}].Item")
                safe_set_cell_rc(ws, row, 5, requested_value, f"Lines[{i}].Requested")
                safe_set_cell_rc(ws, row, 6, issue_value, f"Lines[{i}].Issue")
                safe_set_cell_rc(ws, row, 7, already_issued_value, f"Lines[{i}].AlreadyIssued")
                safe_set_cell_rc(ws, row, 8, balance_value, f"Lines[{i}].Balance")

            # Enforce stable page sizing but preserve template's own print range/layout
            # so logos and decorative elements are not cropped.
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

            temp_xlsx = get_versioned_path(WORK_DIR / f"{processing_path.stem}.xlsx")
            wb.save(temp_xlsx)
            wb.close()
            
            # --- CONVERSION ---
            self.set_status(f"Converting {req_no} to PDF...")
            self.touch_processing_file(processing_path)
            pdf_path = get_versioned_path(OUTPUT_DIR / f"{req_no}.pdf")
            if self.serialize_pdf_export:
                self.set_status(f"Waiting export slot for {req_no}...")

            export_context = self.export_lock if self.serialize_pdf_export else threading.Lock()
            with export_context:
                pythoncom.CoInitialize()
                excel = None
                wb_api = None
                try:
                    excel = win32com.client.DispatchEx("Excel.Application")
                    excel.Visible, excel.DisplayAlerts = False, False
                    wb_api = excel.Workbooks.Open(str(temp_xlsx.absolute()))

                    try:
                        # Targeted Sheet Export
                        ws_to_pdf = wb_api.Worksheets("CollectionVoucher")
                        self.apply_excel_page_setup(ws_to_pdf)
                        ws_to_pdf.ExportAsFixedFormat(0, str(pdf_path.absolute()))
                    except Exception:
                        ws_to_pdf = wb_api.Worksheets(1)
                        self.apply_excel_page_setup(ws_to_pdf)
                        ws_to_pdf.ExportAsFixedFormat(0, str(pdf_path.absolute()))
                finally:
                    try:
                        if wb_api:
                            wb_api.Close(False)
                    except Exception:
                        pass
                    try:
                        if excel:
                            excel.Quit()
                    except Exception:
                        pass
                    try:
                        pythoncom.CoUninitialize()
                    except Exception:
                        pass
            
            # --- CLEANUP ---
            archive_processed_json(processing_path)
            self.clear_retry_count(processing_path)
            if temp_xlsx.exists():
                deleted = False
                for _ in range(5):
                    try:
                        os.remove(temp_xlsx)
                        deleted = True
                        break
                    except PermissionError:
                        time.sleep(0.4)
                    except Exception as cleanup_error:
                        log_message(f"CLEANUP WARNING: Could not delete temp workbook {temp_xlsx.name} - {cleanup_error}")
                        break
                if not deleted and temp_xlsx.exists():
                    log_message(f"CLEANUP WARNING: Temp workbook still present: {temp_xlsx.name}")
            
            # Open the PDF file directly
            try:
                os.startfile(str(pdf_path.absolute()))
                log_message(f"Opened PDF: {pdf_path.name}")
                log_message(f"PDF SAVED: {pdf_path.absolute()}")
            except Exception as e:
                log_message(f"Could not open PDF automatically: {e}")
                # Fallback to opening folder if PDF opening fails
                subprocess.Popen(f'explorer "{OUTPUT_DIR}"')
            
            # Send PDF via email if configured
            email_sent = False
            if self.email_sender.enabled:
                self.set_status(f"Sending email for {req_no}...")
                email_sent = self.email_sender.send_pdf(pdf_path, req_no)
            
            # Update status based on email result
            if email_sent:
                self.set_status(f"Done: {pdf_path.name} (Email sent)", notify=True)
            else:
                self.set_status(f"Done: {pdf_path.name} (Email failed, PDF saved)", notify=True)
                try:
                    subprocess.Popen(f'explorer "{OUTPUT_DIR}"')
                except Exception as e:
                    log_message(f"Could not open output folder after email failure: {e}")
            
        except Exception as e:
            log_message(f"Processing Error: {e}")
            try:
                if processing_path and Path(processing_path).exists():
                    attempt = self.increment_retry_count(processing_path)
                    if attempt >= self.max_process_retries:
                        log_message(f"RECOVERY: Max retries reached for {Path(processing_path).name} ({attempt}/{self.max_process_retries})")
                        self.move_to_failed(processing_path, e, attempt)
                    else:
                        retry_json = Path(processing_path).with_suffix('.json')
                        os.rename(processing_path, retry_json)
                        log_message(f"RECOVERY: Restored {Path(retry_json).name} for retry ({attempt}/{self.max_process_retries})")
            except Exception as recovery_error:
                log_message(f"RECOVERY ERROR: Could not restore JSON for retry - {recovery_error}")
            self.set_status("Error (Check Log)", notify=True)
        finally:
            if processing_path is not None:
                self.unmark_processing_active(processing_path)
            self.set_status("Idle")

    def process_json_from_processing(self, processing_path):
        """Handle recovery of .processing files left from crashes."""
        processing_key = str(processing_path)
        with self.recovery_lock:
            if processing_key in self.recovery_in_progress:
                return
            self.recovery_in_progress.add(processing_key)
        try:
            processing_path = Path(processing_path)
            if not processing_path.exists():
                log_message(f"RECOVERY SKIP: Missing {processing_path.name} (already handled)")
                return

            if is_archival_json_name(processing_path.stem + ".json"):
                log_message(f"RECOVERY SKIP: Ignoring archival artifact {processing_path.name}")
                try:
                    os.remove(processing_path)
                except Exception as cleanup_error:
                    log_message(f"RECOVERY CLEANUP ERROR: Could not remove {processing_path.name} - {cleanup_error}")
                return

            if self.is_processing_active(processing_path):
                log_message(f"RECOVERY SKIP: {processing_path.name} is actively being processed")
                return
            log_message(f"RECOVERING: Processing {processing_path.name}")

            # Rename back to .json and let normal processing handle it.
            # If .json already exists, another worker already re-queued it.
            json_path = processing_path.with_suffix('.json')
            if json_path.exists():
                log_message(f"RECOVERY SKIP: {json_path.name} already exists")
                return

            os.rename(processing_path, json_path)
            self.process_json(str(json_path))
            
        except Exception as e:
            log_message(f"Recovery failed for {processing_path}: {e}")
            try:
                if Path(processing_path).exists() and not self.is_processing_active(processing_path):
                    os.remove(processing_path)
            except Exception:
                pass
        finally:
            with self.recovery_lock:
                self.recovery_in_progress.discard(processing_key)

    def handle_event(self, event):
        # Get the correct path based on event type
        if hasattr(event, 'dest_path'):
            path = event.dest_path
        else:
            path = event.src_path
        
        # Skip if path is empty or None
        if not path:
            console_log(f"EVENT DETECTED: {event.event_type} - EMPTY PATH (OneDrive sync artifact)")
            return
        
        # Log full details for debugging
        console_log(f"EVENT DETECTED: {event.event_type}")
        console_log(f"  Full path: {path}")
        console_log(f"  Is directory: {event.is_directory}")
        
        # Skip OneDrive temporary files
        if "~RF" in path or ".TMP" in path.upper() or ".tmp" in path:
            console_log(f"  --> IGNORED: OneDrive temporary file")
            return
        
        # Check if it's a JSON file
        if path.lower().endswith(".json") and not event.is_directory:
            if self.should_skip_json_artifact(path):
                console_log(f"  --> IGNORED: {Path(path).name} (archival artifact)")
                return
            console_log(f"  --> PROCESSING: {Path(path).name}")
            # Add small delay to ensure OneDrive sync is complete
            time.sleep(0.5)
            threading.Thread(target=self.process_json, args=(path,), daemon=True).start()
        else:
            console_log(f"  --> IGNORED: {Path(path).name} (not .json or is directory)")

    def poll_for_json_files(self):
        """Fallback scanner for environments where file-system events may be missed."""
        while not self.stop_event.is_set():
            try:
                now = time.time()
                pending_json = []
                stale_processing = []

                for entry in JSON_DIR.iterdir():
                    if not entry.is_file():
                        continue
                    suffix = entry.suffix.lower()
                    if suffix == ".json":
                        if self.should_skip_json_artifact(entry):
                            continue
                        pending_json.append(entry)
                    elif suffix == ".processing":
                        if self.is_processing_active(entry):
                            continue
                        age_seconds = now - entry.stat().st_mtime
                        if age_seconds >= self.processing_stale_seconds:
                            stale_processing.append(entry)

                if pending_json:
                    console_log(f"FALLBACK SCAN: Found {len(pending_json)} pending JSON file(s)")
                    for json_file in pending_json:
                        threading.Thread(target=self.process_json, args=(str(json_file),), daemon=True).start()

                if stale_processing:
                    console_log(f"FALLBACK SCAN: Recovering {len(stale_processing)} stale .processing file(s)")
                    for processing_file in stale_processing:
                        threading.Thread(target=self.process_json_from_processing, args=(str(processing_file),), daemon=True).start()
            except Exception as e:
                log_message(f"Fallback scan error: {e}")

            self.stop_event.wait(5)

    def run(self):
        initialize_folders()
        self.prevent_system_sleep()
        self.set_status("Monitoring (Sleep blocked)")
        
        try:
            # Display monitoring info
            console_log("="*60)
            console_log(f"MONITORING FOLDER: {JSON_DIR.absolute()}")
            console_log(f"PDF OUTPUT FOLDER: {OUTPUT_DIR.absolute()}")
            console_log(f"WORK FOLDER: {WORK_DIR.absolute()}")
            console_log(f"Monitoring folder exists: {JSON_DIR.exists()}")
            console_log(f"Monitoring folder is accessible: {os.access(JSON_DIR, os.R_OK)}")
            console_log("="*60)

            # Requeue the two most recent failed requests on startup.
            self.requeue_recent_failed(count=2)

            # Snapshot backlog before processing to avoid startup races
            existing_json = list(JSON_DIR.glob("*.json"))
            existing_json = [entry for entry in existing_json if not self.should_skip_json_artifact(entry)]
            processing_files = list(JSON_DIR.glob("*.processing"))

            # Process crash leftovers first
            if processing_files:
                console_log(f"Found {len(processing_files)} leftover .processing file(s)")
                for processing_file in processing_files:
                    log_message(f"RECOVERING: Found leftover {processing_file.name} from previous crash")
                    threading.Thread(target=self.process_json_from_processing, args=(str(processing_file),), daemon=True).start()

            # Then process JSON backlog
            if existing_json:
                console_log(f"Found {len(existing_json)} existing JSON file(s) to process")
                for json_file in existing_json:
                    console_log(f"  - {json_file.name}")
                    threading.Thread(target=self.process_json, args=(str(json_file),), daemon=True).start()
            else:
                console_log("No existing JSON files found in backlog")
            
            # 1. Start Watcher
            event_handler = PatternMatchingEventHandler(patterns=["*.json"], ignore_directories=True)
            event_handler.on_created = self.handle_event
            event_handler.on_moved = self.handle_event
            event_handler.on_modified = self.handle_event  # Also watch for modifications (OneDrive)
            
            self.observer = Observer()
            self.observer.schedule(event_handler, str(JSON_DIR), recursive=False)
            self.observer.start()
            console_log("File watcher started successfully!")

            # Start fallback polling thread to catch files if watchdog events are missed
            self.poll_thread = threading.Thread(target=self.poll_for_json_files, daemon=True)
            self.poll_thread.start()
            console_log("Fallback scanner started (5s interval)")

            # 3. Tray Icon
            img = Image.new('RGB', (64, 64), (31, 117, 254))
            draw = ImageDraw.Draw(img)
            draw.text((15, 25), "PLZ", fill="white")
            menu = Menu(
                MenuItem("Check Status", lambda i, item: i.notify(self.status, title="PLZ CV Engine")),
                MenuItem("Open PDF Folder", lambda: subprocess.Popen(f'explorer "{OUTPUT_DIR}"')),
                MenuItem("Exit", self.quit_action)
            )
            self.icon = Icon("PLZ", img, "PLZ CV Engine", menu)
            self.icon.run()
        finally:
            self.restore_system_sleep()

    def quit_action(self, icon):
        self.stop_event.set()
        try:
            if self.observer:
                self.observer.stop()
                self.observer.join()

            if self.poll_thread and self.poll_thread.is_alive():
                self.poll_thread.join(timeout=2)

            icon.stop()
        finally:
            self.restore_system_sleep()

if __name__ == "__main__":
    try:
        runtime_mode = "frozen-exe" if getattr(sys, 'frozen', False) else "python-script"
        runtime_entry = sys.executable if getattr(sys, 'frozen', False) else __file__
        console_log(f"RUNTIME: mode={runtime_mode} entry={runtime_entry}")
    except Exception:
        pass
    SplashScreen().show()
    CVAutomator().run()